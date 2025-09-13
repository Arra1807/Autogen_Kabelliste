# creating Cablelist
import pandas as pd
import numpy as np
import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
#Defining a path
path = 'C:\\Users\\shado\\Documents\\Bacheloroppgave\\kabelliste.xlsx'

try:
    #Open master signalliste (default liste)
    default_frame = pd.read_excel('Master-signaliste.xlsx', sheet_name= "Objektliste")
    # Open the Excel sheet 2
    df2 = pd.read_excel('signalliste.xlsx', sheet_name="Ark2")
    # Open excel sheet 1
    df1 = pd.read_excel('signalliste.xlsx', sheet_name='Ark1')
except ValueError as ve:
    if str(ve).startswith(f"Worksheet named: {'sheet_name'} not found"):
        print(f"Worksheet {'sheet_name'} not found in the specific Excel file")
    else:
        print(f"ValueError occurred: {ve}")
except Exception as e:
    print(f"Error occurred: {e}")
    
#Droping unneccesarry columns
df2 = df2.drop('Systemkode', axis=1)
df2 = df2.drop('For belysningskabel mellom armaturer',axis =1)
df2 = df2.drop('Unnamed: 12', axis =1)
df2 = df2.drop('Unnamed: 4', axis = 1)
df2 = df2.drop(['Unnamed: 2','Unnamed: 14','Unnamed: 16'], axis = 1)
df2 = df2.drop(['Unnamed: 17','Unnamed: 18','Unnamed: 19'], axis = 1)
df2 = df2.drop(['Komponentkode','Unnamed: 6', 'Lokaliseringskode', 'Unnamed: 1'], axis = 1)

#Creating index  column
df2.insert(0,"Unnamed: 0", range(0, len(df2)))    
df2.iat[0,0] = 'Index'

#Changing name of column 2 to Kabel fra
df2.rename(columns={'Tag for SOS/Fordeling/felt etc. som forsyner kablen/utstyret direkte ELLER Armaturer/kobl.boks med utfylt segmentlengde': 'Kabel Fra'}, inplace=True)
df2.head()

#Creating klemmeliste column
df2.insert(3,"Unnamed: 1", "")
df2.iat[0,3] = 'Klemmeliste'

#Creating kabel column
df2.insert(4,"Kabel","")
for idx, rows in df2.iterrows():
    rows[4] = f"W{100+idx}"
df2.iat[0,4] = 'KabelType nr.'

#Modifying Kabel til column
df2.rename(columns={'Unnamed: 10': 'Kabel Til'}, inplace=True)
df2.iat[0,5] = 'Objektnavn'

#Creating Beskrivelse column
df2.insert(7, "Unnamed: 4", " ")
start = 1
liste_objekt = []
#default_frame.dropna(inplace = True)
for idx_df2, objekt_df2 in df2.iloc[start:].iterrows():
    objekt_navn = objekt_df2[5][5:9]
    liste_objekt.append(objekt_navn)

objektbeskrivelse = {}
start = 1
for idx_default, row_default in default_frame.iloc[start:].iterrows():
    if row_default[1] in liste_objekt:
        # Check if the value is not NaN before assigning it
        if not pd.isnull(row_default[4]):
            objektbeskrivelse[row_default[1]] = row_default[4]

start =1
for idx, row in df2.iloc[start:].iterrows():
    objekt_kode = row[5][5:9]
    if objekt_kode in objektbeskrivelse:
        df2.at[idx, 'Unnamed: 4']= objektbeskrivelse[objekt_kode]    
    elif row[5][5:8] == 'ITV':
        df2.at[idx, 'Unnamed: 4'] = 'Kamera'
# Update the first row of the 'Unnamed: 4' column
df2.iat[0, 7] = 'Beskrivelse'

#Creating SignalType column
df2.insert(8, "Unnamed: 12", " ")
df1.replace(np.nan, 0, inplace = True)

for idx, rows in df1.iterrows():   
    if df1.loc[idx,"DO"] == 1 or df1.loc[idx,"DI"] ==1: 
        df2["Unnamed: 12"] = 'Digital'
    elif df1.loc[idx,'AI'] == 1 or df1.loc[idx, 'AO'] == 1:
        df2['Unnamed: 12'] = "4-20mA"

#Checking for any opc-tags that does not exist
opc_tags_df1 = set(df1['OPC-tag'])
opc_tags_df2 = set(df2['Kabel Til'])

# Find the OPC-tags that exist in df2 but not in df1
missing_opc_tags = opc_tags_df2 - opc_tags_df1

# Update 'Unnamed: 11' column in df2 for missing OPC-tags
df2.loc[df2['Kabel Til'].isin(missing_opc_tags), 'Unnamed: 12'] = 'Eksisterer ikke'    

for idx, rows in df2.iterrows():    
    if df2['Kabel Til'][idx][5:8] == 'ITV':
        df2.at[idx, 'Unnamed: 12'] = "No Signal"
df2.iat[0,8] = 'SignalType'

#Modifying Lengde column
start = 1
for idx, row in df2.iloc[start:].iterrows():
    lengde_fra_profil = int(row[1]) #Converting string elements i fra profil to integer
    lengde_til_profil = row[6] #Assigning row 6 as lengde_1 
    total_lengde = lengde_fra_profil-lengde_til_profil #Subtracting Lengde and len
    df2.at[idx, 'Unnamed: 13'] = total_lengde
    if lengde_fra_profil < lengde_til_profil:
        df2.at[idx, 'Unnamed: 13'] = abs(total_lengde)
    
    
#Creating Ledninger column 
df2.insert(11, 'Unnamed: 16', " ")
kabel_til_set = set(df2['Kabel Til'])
kabel_til_set.remove('Objektnavn')
indices_to_drop = []
# Iterate over the rows of df1
for idx, row in df1.iterrows():
    # Check if the 'OPC-tag' value is not in kabel_til_set
    if row['OPC-tag'] not in kabel_til_set:
        # Add index to the list of indices to drop
        indices_to_drop.append(idx)
# Drop rows based on the indices in the list
df1.drop(indices_to_drop, inplace=True)
resultat = df1.groupby(['OPC-tag'])['OPC-tag'].count().reset_index(name = 'Counts')
df3 = pd.DataFrame(resultat)
df3.index = df3.index+1
start = 1
for idx, row in df2.iloc[start:].iterrows():
    df2['Unnamed: 16'] = df3['Counts']
    if 'NAN' in df2['Unnamed: 16']:
        df2['Unnamed: 16'].fillna(0, inplace = True)
df2['Unnamed: 16'].replace(np.nan, 0, inplace= True)
df2.iat[0,11] = 'Ledninger' 

#Modifying KabelType column 
start = 1
for idx, row in df2.iloc[start:].iterrows():
    antall_ledninger = row[11]
    df2.at[idx, 'Unnamed: 15'] = f"{row[10][0:len(row[10])-9]} {antall_ledninger+2}x2,5 Cu"      
    if pd.notna(df2['Kabel Til'][idx]):
        if df2['Kabel Til'][idx][5:8] == 'ITV':
            df2.at[idx, 'Unnamed: 15'] =f'{row[10][0:len(row[10])-9]} 2x2,5 Cu'
        elif df2['Kabel Til'][idx][5:8] == 'ITV':
            df2.at[idx, 'Unnamed: 15'] ='1x Ethernet'
df2.head()
df2.iat[0,10] = 'Kabeltype'

#Creating Kabelmerke column
df2.insert(12, "Unnamed: 17", "")
for idx, rows in df2.iterrows():
    rows['Unnamed: 17'] = f"+{rows['Kabel Fra']}.{rows['Unnamed: 1']}={rows['Kabel Til']}-{rows['Kabel']}"
df2.iat[0,12] = 'Kabelmerke'

#start = 1
#for idx, rows in df2.iloc[start:].iterrows():
#    if rows[5][5:8] == 'ITV':
        
    
#Funksjon for Excel formattering
def save_to_excel(df, file_path):
    try:
      # Convert DataFrame to Excel
      df.to_excel(file_path, index=False)  # Set index=False if you don't want to include the DataFrame index in the Excel file
      print(f"DataFrame successfully saved to {file_path}")

      # Load the workbook
      workbook = openpyxl.load_workbook(file_path)
      sheet = workbook.active  # Get the active sheet
      
      #Remove the unwanted column names
      coordinates_to_exclude = ['A1','B1','D1','G1','H1','I1','J1','K1', 'L1','M1'] #Listing the unwanted named by their coordinates
      for cell_coordinate in coordinates_to_exclude: #For loop to iterate the list with unwanted names
          sheet[cell_coordinate].value = None
          
      #Slå sammen de 3 øverste cellene, 1 celle mellomrom, slå sammen 3 nye celler  
      sheet.merge_cells(start_row=1, start_column=2, end_row=1, end_column=4)
      sheet.cell(row=1, column=2).value = 'Kabel Fra'
      sheet.merge_cells(start_row=1, start_column = 6, end_row = 1, end_column= 8)    
      
      # Fjern celle rammer for de øverste 13 cellene i regnearket
      for rad in sheet.iter_rows(min_row =1,max_row= 14, min_col= 1, max_col = 1):
          for celler in rad:
              celler.border = None
    
      # Apply cell formatting
      for col in range(2, 9):  # Green color for column 2,3,5,6 
          sheet.cell(row=1, column=col).fill = PatternFill(start_color="008000", end_color="008000", fill_type="solid")
      sheet['E1'].fill = PatternFill(start_color="404040", end_color="404040", fill_type="solid") #Color for Kabel column
      for col in range(1, 14):
          sheet.cell(row=2, column=col).fill = PatternFill(start_color="0343DF", end_color="0343DF", fill_type="solid") 
      
      #Color the KabelType column Yellow
      column_idx = df2.columns.get_loc('Unnamed: 15') + 1 #Getting the column
      fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid") #Choosing the colour by checking the color codes
      for row_idx in range(3, len(df)+2):  # Start from row 2 (assuming header is in row 1)
            sheet.cell(row=row_idx, column=column_idx).fill = fill 
            
      #Forstørrer cellen, med hensyn til lengden 
      for col_idx, col in enumerate(df.columns, start=1): 
          max_length = max(df[col].astype(str).map(len).max(), len(col))
          adjusted_width = (max_length + 2) * 1.2  # JUstert bredde basert på den lengste strengen i kolonnen
          sheet.column_dimensions[get_column_letter(col_idx)].width = adjusted_width  
        
     #Teksten midtstilles, med hensyn til lengden på teksten 
      for row in sheet.iter_rows():
            for cell in row:
                alignment = Alignment(horizontal='center', vertical='center')
                cell.alignment = alignment
      
    # Save the workbook
      workbook.save(file_path)
      print("Cell formatting applied successfully")
    except Exception as e:
      print(f"Error occurred: {e}")

save_to_excel(df2, path) #Lagres som excel fil



